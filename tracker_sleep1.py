import pandas as pd
import matplotlib.pyplot as plt
from datetime import datetime
import os
import msvcrt
from openpyxl.styles import Font, Border, Side, Alignment

# Настройки путей
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
FILE_NAME = os.path.join(BASE_DIR, 'sleep_log.xlsx')
CHART_NAME = os.path.join(BASE_DIR, 'sleep_chart.png')

def smart_input(prompt):
    print(f"\n[ {prompt} ]")
    template = "__.__.____ __:__"
    input_str = ""
    idx_map = [0, 1, 3, 4, 6, 7, 8, 9, 11, 12, 14, 15]
    
    while len(input_str) < 12:
        display = list(template)
        for i, char in enumerate(input_str):
            display[idx_map[i]] = char
        print(f"\r│ {''.join(display)}", end='', flush=True)
        
        char = msvcrt.getch()
        if char == b'\x08': # Backspace
            input_str = input_str[:-1]
        elif char.isdigit():
            input_str += char.decode('utf-8')
            
    final_date = f"{input_str[:2]}.{input_str[2:4]}.{input_str[4:8]} {input_str[8:10]}:{input_str[10:]}"
    print(f"\r│ {final_date} ✅")
    return datetime.strptime(final_date, '%d.%m.%Y %H:%M')

def format_excel(file_path):
    from openpyxl import load_workbook
    wb = load_workbook(file_path)
    header_font = Font(size=12, bold=True)
    bold_font = Font(bold=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))

    ws = wb['Данные']
    for cell in ws[1]: # Заголовки
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=7):
        for cell in row:
            cell.border = thin_border
            if cell.column < 7: cell.alignment = Alignment(horizontal='center')

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['G'].width = 45

    ws_stat = wb['Аналитика']
    ws_stat.column_dimensions['A'].width = 35
    for row in ws_stat.iter_rows():
        row[0].font = bold_font
        for cell in row: cell.border = thin_border
    wb.save(file_path)

def process_and_save(df):
    """Сортирует, пересчитывает и сохраняет всё в файл."""
    # 1. Сортировка и пересчет
    df['Лег спать'] = pd.to_datetime(df['Лег спать'], dayfirst=True)
    df['Проснулся'] = pd.to_datetime(df['Проснулся'], dayfirst=True)
    df = df.sort_values(by='Лег спать').reset_index(drop=True)

    awakes, shifts, totals = [0.0]*len(df), [0.0]*len(df), [0.0]*len(df)
    for i in range(1, len(df)):
        p_wake, p_bed = df.loc[i-1, 'Проснулся'], df.loc[i-1, 'Лег спать']
        c_bed, c_sleep = df.loc[i, 'Лег спать'], df.loc[i, 'Сон']
        gap = (c_bed - p_wake).total_seconds() / 3600
        if 0 < gap < 48:
            awakes[i] = round(gap, 2)
            shifts[i] = round((c_bed - p_bed).total_seconds() / 3600 - 24, 2)
            totals[i] = round(c_sleep + awakes[i], 2)

    df['Бодрствование'], df['Сдвиг'], df['Сутки'] = awakes, shifts, totals
    df['Лег спать'] = df['Лег спать'].dt.strftime('%d.%m.%Y %H:%M')
    df['Проснулся'] = df['Проснулся'].dt.strftime('%d.%m.%Y %H:%M')

        # 2. Расширенная статистика
    valid = df[df['Бодрствование'] > 0]
    stats_data = [
        ("Средний сон", df['Сон'].mean()), 
        ("Минимальный сон", df['Сон'].min()), 
        ("Максимальный сон", df['Сон'].max()),
        ("Среднее бодрствование", valid['Бодрствование'].mean() if not valid.empty else 0),
        ("Минимальное бодрствование", valid['Бодрствование'].min() if not valid.empty else 0),
        ("Максимальное бодрствование", valid['Бодрствование'].max() if not valid.empty else 0),
        ("Средние сутки", valid['Сутки'].mean() if not valid.empty else 0),
        ("Минимальные сутки", valid['Сутки'].min() if not valid.empty else 0),
        ("Максимальные сутки", valid['Сутки'].max() if not valid.empty else 0),
        ("Минимальный сдвиг (ранее)", valid['Сдвиг'].min() if not valid.empty else 0),
        ("Максимальный сдвиг (позже)", valid['Сдвиг'].max() if not valid.empty else 0),
        ("Всего периодов сна", len(df)), 
        ("Всего часов сна за всё время", df['Сон'].sum())
    ]
    
    stats_df = pd.DataFrame(stats_data)
    # Округление всех чисел до 1 знака
    stats_df[1] = stats_df[1].apply(lambda x: round(x, 1) if isinstance(x, (float, int)) else x)


    # 3. Запись
    with pd.ExcelWriter(FILE_NAME, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Данные', index=False)
        stats_df.to_excel(writer, sheet_name='Аналитика', index=False, header=False)
    
    format_excel(FILE_NAME)

    # 4. График
    plt.figure(figsize=(10, 5))
    plt.plot(df.index, df['Сон'], 'b-o', label='Сон')
    plt.plot(df.index, df['Бодрствование'].replace(0, float('nan')), 'r-s', label='Бодрствование')
    plt.axhline(0, color='gray', linestyle='--', alpha=0.5)
    plt.title('График сна')
    plt.legend(); plt.grid(alpha=0.3); plt.savefig(CHART_NAME); plt.close()

def main():
    print("=== УЧЕТ СНА (Закройте окно для выхода) ===")
    while True:
        try:
            if os.path.exists(FILE_NAME):
                df = pd.read_excel(FILE_NAME, sheet_name='Данные')
            else:
                df = pd.DataFrame(columns=['Лег спать', 'Проснулся', 'Сон', 'Бодрствование', 'Сдвиг', 'Сутки', 'Комментарий'])

            bed_t = smart_input("Когда легли?")
            wake_t = smart_input("Когда проснулись?")
            comment = input("\n[ Комментарий ] >> ")
            
            sleep_dur = round((wake_t - bed_t).total_seconds() / 3600, 2)
            new_row = {'Лег спать': bed_t.strftime('%d.%m.%Y %H:%M'), 
                       'Проснулся': wake_t.strftime('%d.%m.%Y %H:%M'), 
                       'Сон': sleep_dur, 'Комментарий': comment}
            
            df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
            
            process_and_save(df)
            print("\n✅ Сохранено! Нажмите Enter для новой строки или закройте окно.")
            input()

        except Exception as e:
            print(f"\n🚨 Ошибка: {e}")
            input("Нажмите Enter, чтобы попробовать снова...")

if __name__ == "__main__":
    main()
