import pandas as pd
import matplotlib.pyplot as plt
from datetime import datetime, timedelta
import os
import msvcrt
from openpyxl.styles import Font, Border, Side, Alignment
import sqlite3
import shutil
import tempfile
import subprocess

# ---------- прежние функции (ручной ввод и оформление) ----------
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
FILE_NAME = os.path.join(BASE_DIR, 'sleep_log.xlsx')
CHART_NAME = os.path.join(BASE_DIR, 'sleep_chart.png')

def smart_input_time(prompt):
    """Ввод времени в формате ЧЧ:ММ"""
    print(f"\n[ {prompt} ]")
    template = "__:__"
    input_str = ""
    idx_map = [0, 1, 3, 4]
    while len(input_str) < 4:
        display = list(template)
        for i, char in enumerate(input_str):
            display[idx_map[i]] = char
        print(f"\r│ {''.join(display)}", end='', flush=True)
        char = msvcrt.getch()
        if char == b'\x08':            # Backspace
            input_str = input_str[:-1]
        elif char.isdigit():
            input_str += char.decode('utf-8')
    try:
        h, m = int(input_str[:2]), int(input_str[2:])
        if 0 <= h <= 23 and 0 <= m <= 59:
            time_str = f"{input_str[:2]}:{input_str[2:]}"
            print(f"\r│ {time_str} ✅")
            return time_str
        else:
            print(f"\r│ {input_str[:2]}:{input_str[2:]} ❌ Некорректное время")
            return smart_input_time(prompt)
    except:
        print(f"\r│ {input_str[:2]}:{input_str[2:]} ❌ Ошибка")
        return smart_input_time(prompt)

def smart_input_date(prompt):
    """Ввод даты ДД.ММ.ГГГГ, при вводе только дня и месяца + Enter год = текущий"""
    print(f"\n[ {prompt} ]")
    template = "__.__.____"
    curr_year = str(datetime.now().year)
    input_str = ""
    idx_map = [0, 1, 3, 4, 6, 7, 8, 9]
    while True:
        display = list(template)
        for i, char in enumerate(input_str):
            display[idx_map[i]] = char
        print(f"\r│ {''.join(display)}", end='', flush=True)
        char = msvcrt.getch()
        if char == b'\x08':
            input_str = input_str[:-1]
        elif char == b'\r':                     # Enter
            if len(input_str) == 4:
                input_str += curr_year
                break
            elif len(input_str) == 8:
                break
        elif char.isdigit() and len(input_str) < 8:
            input_str += char.decode('utf-8')
    date_str = f"{input_str[:2]}.{input_str[2:4]}.{input_str[4:]}"
    try:
        datetime.strptime(date_str, '%d.%m.%Y')
        print(f"\r│ {date_str} ✅")
        return date_str
    except:
        print(f"\r│ {date_str} ❌ Ошибка в дате")
        return smart_input_date(prompt)

def format_excel(file_path):
    """Красивое оформление Excel"""
    from openpyxl import load_workbook
    wb = load_workbook(file_path)
    header_font = Font(size=12, bold=True)
    bold_font = Font(bold=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    if 'Данные' in wb.sheetnames:
        ws = wb['Данные']
        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=7):
            for cell in row:
                cell.border = thin_border
                if cell.column < 7:
                    cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['G'].width = 45

    if 'Аналитика' in wb.sheetnames:
        ws_stat = wb['Аналитика']
        ws_stat.column_dimensions['A'].width = 35
        for row in ws_stat.iter_rows():
            row[0].font = bold_font
            for cell in row:
                cell.border = thin_border
    wb.save(file_path)

def process_and_save(df):
    """Сортировка, расчёт показателей, статистика, Excel и график"""
    df['Лег спать'] = pd.to_datetime(df['Лег спать'], dayfirst=True)
    df['Проснулся'] = pd.to_datetime(df['Проснулся'], dayfirst=True)
    df = df.sort_values(by='Лег спать').reset_index(drop=True)

    awakes, shifts, totals = [0.0]*len(df), [0.0]*len(df), [0.0]*len(df)
    for i in range(len(df)):
        sleep_dur = (df.loc[i, 'Проснулся'] - df.loc[i, 'Лег спать']).total_seconds() / 3600
        df.loc[i, 'Сон'] = round(sleep_dur, 2)
        if i > 0:
            p_wake = df.loc[i-1, 'Проснулся']
            p_bed = df.loc[i-1, 'Лег спать']
            c_bed = df.loc[i, 'Лег спать']
            gap = (c_bed - p_wake).total_seconds() / 3600
            if 0 < gap < 200:
                awakes[i] = round(gap, 2)
                shifts[i] = round((c_bed - p_bed).total_seconds() / 3600 - 24, 2)
                totals[i] = round(sleep_dur + awakes[i], 2)

    df['Бодрствование'] = awakes
    df['Сдвиг'] = shifts
    df['Сутки'] = totals

    valid = df[df['Бодрствование'] > 0]
    stats_data = [
        ("Средний сон", df['Сон'].mean()),
        ("Минимальный сон", df['Сон'].min()),
        ("Максимальный сон", df['Сон'].max()),
        ("Среднее бодрствование", valid['Бодрствование'].mean() if not valid.empty else 0),
        ("Минимальное бодрствование", valid['Бодрствование'].min() if not valid.empty else 0),
        ("Максимальное бодрствование", valid['Бодрствование'].max() if not valid.empty else 0),
        ("Средние сутки", valid['Сутки'].mean() if not valid.empty else 0),
        ("Минимальные сутки", valid['Сутки'].min() if not valid.empty else 0),
        ("Максимальные сутки", valid['Сутки'].max() if not valid.empty else 0),
        ("Минимальный сдвиг (ранее)", valid['Сдвиг'].min() if not valid.empty else 0),
        ("Максимальный сдвиг (позже)", valid['Сдвиг'].max() if not valid.empty else 0),
        ("Всего периодов сна", len(df)),
        ("Всего часов сна за всё время", df['Сон'].sum())
    ]
    stats_df = pd.DataFrame(stats_data)
    stats_df[1] = stats_df[1].apply(lambda x: round(x, 1) if isinstance(x, (float, int)) else x)

    df_excel = df.copy()
    df_excel['Лег спать'] = df_excel['Лег спать'].dt.strftime('%d.%m.%Y %H:%M')
    df_excel['Проснулся'] = df_excel['Проснулся'].dt.strftime('%d.%m.%Y %H:%M')

    with pd.ExcelWriter(FILE_NAME, engine='openpyxl') as writer:
        df_excel.to_excel(writer, sheet_name='Данные', index=False)
        stats_df.to_excel(writer, sheet_name='Аналитика', index=False, header=False)

    format_excel(FILE_NAME)

    plt.figure(figsize=(10, 5))
    plt.plot(df.index, df['Сон'], 'b-o', label='Сон')
    plt.plot(df.index, df['Бодрствование'].replace(0, float('nan')), 'r-s', label='Бодрствование')
    plt.title('История сна и активности')
    plt.legend()
    plt.grid(alpha=0.3)
    plt.savefig(CHART_NAME)
    plt.close()

# ---------- автоматический сбор данных ----------
def get_chrome_visits():
    """Извлекает времена посещений из истории Chrome."""
    visits = []
    chrome_dir = os.path.expandvars(r'%LOCALAPPDATA%\Google\Chrome\User Data')
    if not os.path.isdir(chrome_dir):
        return visits

    for profile in ['Default', 'Profile 1', 'Profile 2']:
        history_path = os.path.join(chrome_dir, profile, 'History')
        if not os.path.isfile(history_path):
            continue
        try:
            # Пробуем открыть напрямую
            conn = sqlite3.connect(history_path)
            _read_chrome_timestamps(conn, visits)
            conn.close()
        except Exception:
            # Файл заблокирован – копируем во временный
            try:
                tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.sqlite')
                tmp.close()
                shutil.copy2(history_path, tmp.name)
                conn = sqlite3.connect(tmp.name)
                _read_chrome_timestamps(conn, visits)
                conn.close()
                os.unlink(tmp.name)
            except Exception:
                pass
    return visits

def _read_chrome_timestamps(conn, out_list):
    """Читает last_visit_time из открытого соединения и добавляет datetime в out_list."""
    cursor = conn.cursor()
    cursor.execute("SELECT last_visit_time FROM urls WHERE last_visit_time > 0")
    for row in cursor:
        t = row[0]
        # Chrome time: микросекунды с 1601-01-01 → секунды Unix → локальное время
        try:
            unix_ts = (t / 1_000_000) - 11644473600
            if unix_ts > 0:
                dt = datetime.fromtimestamp(unix_ts)
                out_list.append(dt)
        except:
            pass

def get_windows_power_events():
    """Получает времена событий включения/выключения из журнала System (ID 6005/6006)."""
    events = []
    try:
        cmd = [
            'powershell', '-Command',
            "Get-WinEvent -FilterHashtable @{LogName='System'; ID=6005,6006} -MaxEvents 500 "
            "| ForEach-Object { $_.TimeCreated.ToString('yyyy-MM-dd HH:mm:ss') }"
        ]
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=30)
        if result.returncode == 0:
            for line in result.stdout.splitlines():
                line = line.strip()
                if line:
                    try:
                        dt = datetime.strptime(line, '%Y-%m-%d %H:%M:%S')
                        events.append(dt)
                    except:
                        pass
    except Exception as e:
        print(f"Ошибка получения событий Windows: {e}")
    return events

def find_sleep_periods(activity_times, min_gap_h=3.0, max_gap_h=24.0):
    """
    Находит непрерывные промежутки без активности длительностью
    от min_gap_h до max_gap_h часов и превращает их в периоды сна.
    Возвращает список кортежей (bed_time, wake_time).
    """
    if len(activity_times) < 2:
        return []

    # Убираем дубликаты и сортируем
    unique = sorted(set(activity_times))
    periods = []
    for i in range(len(unique) - 1):
        t1 = unique[i]
        t2 = unique[i + 1]
        gap_hours = (t2 - t1).total_seconds() / 3600.0
        if min_gap_h < gap_hours <= max_gap_h:
            bed = t1 + timedelta(minutes=30)      # время засыпания
            wake = t2 - timedelta(hours=1)        # время подъёма
            if bed < wake:
                periods.append((bed, wake))
    return periods

# ---------- главный блок ----------
def main():
    print("=== АВТОМАТИЧЕСКИЙ УЧЕТ СНА (по логам активности) ===\n")

    # 1. Загружаем существующий лог
    if os.path.exists(FILE_NAME):
        df = pd.read_excel(FILE_NAME, sheet_name='Данные')
    else:
        df = pd.DataFrame(columns=[
            'Лег спать', 'Проснулся', 'Сон', 'Бодрствование',
            'Сдвиг', 'Сутки', 'Комментарий'
        ])

    # Определяем дату, после которой ищем новые записи
    if df.empty:
        max_date = None
    else:
        # Приводим к datetime для надёжности
        df['Проснулся_dt'] = pd.to_datetime(df['Проснулся'], dayfirst=True)
        max_date = df['Проснулся_dt'].max()
        df.drop(columns=['Проснулся_dt'], inplace=True, errors='ignore')

    # 2. Собираем события активности
    print("Сбор истории Chrome...")
    chrome_events = get_chrome_visits()
    print(f"Найдено посещений Chrome: {len(chrome_events)}")

    print("Получение событий включения/выключения Windows...")
    power_events = get_windows_power_events()
    print(f"Найдено событий питания: {len(power_events)}")

    all_events = chrome_events + power_events
    if not all_events:
        print("\nНе удалось найти никаких данных об активности. Программа завершена.")
        input("Нажмите Enter для выхода...")
        return

    # 3. Ищем периоды сна
    sleep_periods = find_sleep_periods(all_events)
    print(f"Обнаружено периодов сна (перерыв >3ч и ≤16ч): {len(sleep_periods)}")

    # 4. Отбираем новые (после последней записи)
    if max_date is not None:
        new_periods = [p for p in sleep_periods if p[0] > max_date]
    else:
        new_periods = sleep_periods

    # 5. Вывод результатов
    # Сохранённый лог (последние 5 записей)
    if not df.empty:
        print("\n1. Сохраненный лог")
        for _, row in df.tail(5).iterrows():
            bed_str = row['Лег спать'] if isinstance(row['Лег спать'], str) else pd.to_datetime(row['Лег спать']).strftime('%d.%m.%Y %H:%M')
            wake_str = row['Проснулся'] if isinstance(row['Проснулся'], str) else pd.to_datetime(row['Проснулся']).strftime('%d.%m.%Y %H:%M')
            sleep_val = row['Сон'] if not pd.isna(row['Сон']) else 0
            awake_val = row['Бодрствование'] if not pd.isna(row['Бодрствование']) else 0
            print(f"{bed_str} - {wake_str}   сон {round(sleep_val,2)} бодрствование {round(awake_val,2)}")
    else:
        print("\n1. Сохраненный лог пуст")

    print("\n2. Новые даты.")
    if not new_periods:
        print("Нет новых записей для добавления.")
        input("Нажмите Enter для выхода...")
        return

    prev_wake = max_date
    for i, (bed, wake) in enumerate(new_periods, 1):
        sleep_dur = round((wake - bed).total_seconds() / 3600, 2)
        if prev_wake is not None:
            awake_gap = round((bed - prev_wake).total_seconds() / 3600, 2)
            awake_str = f"бодрствование {awake_gap}"
        else:
            awake_str = "бодрствование -"
        print(f"{i}. {bed.strftime('%d.%m.%Y %H:%M')} - {wake.strftime('%d.%m.%Y %H:%M')}   сон {sleep_dur} {awake_str}")
        prev_wake = wake

    # 6. Подтверждение
    print("\nВсе даты верны, сохраняем?\n1. Да\n2. Нет")
    choice = input(">> ").strip()

    if choice == '2':
        try:
            nums = input("Введите через пробел номера ошибочных дат: ").strip()
            remove_indices = sorted([int(x) for x in nums.split()], reverse=True)
            for idx in remove_indices:
                if 1 <= idx <= len(new_periods):
                    del new_periods[idx - 1]
        except:
            print("Ошибка ввода. Новые даты не изменены.")

    if new_periods:
        # Добавляем оставшиеся записи в общий DataFrame
        new_rows = []
        for bed, wake in new_periods:
            new_rows.append({
                'Лег спать': bed.strftime('%d.%m.%Y %H:%M'),
                'Проснулся': wake.strftime('%d.%m.%Y %H:%M'),
                'Комментарий': ''
            })
        new_df = pd.DataFrame(new_rows)
        df = pd.concat([df, new_df], ignore_index=True)
        process_and_save(df)
        print("\n✅ Лог сна обновлён!")
    else:
        print("Нет новых дат для сохранения.")

    input("\nНажмите Enter, чтобы закрыть окно...")

if __name__ == "__main__":
    main()